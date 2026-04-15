from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

# Definir estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=10)
category_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
category_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Dados do mapeamento
campos = [
    {
        "categoria": "1. Controle e Relacionamento",
        "campos": [
            ("id", "INTEGER", "Identificador único da tabela (chave primária)"),
            ("arquivo_id", "INTEGER", "FK para efd_arquivo - Referência ao arquivo EFD processado"),
            ("efd_resumo_id", "INTEGER", "FK para efd_resumo - Referência ao registro C100 (documento fiscal)"),
            ("cod_reg", "INTEGER", "Código do Registro = 170 (constante para todos os registros)"),
        ]
    },
    {
        "categoria": "2. Informações de Operação",
        "campos": [
            ("ind_oper", "INTEGER", "Indicador da Operação (0=Entrada, 1=Saída, 2=Transferência)"),
            ("ind_emit", "INTEGER", "Indicador do Emitente (0=Própria empresa, 1=Terceiro)"),
            ("dt_doc", "TIMESTAMP", "Data de emissão do documento fiscal"),
            ("dt_e_s", "TIMESTAMP", "Data de entrada/saída da mercadoria"),
        ]
    },
    {
        "categoria": "3. Identificação do Documento Fiscal (Cabeçalho C100)",
        "campos": [
            ("cod_mod", "VARCHAR(2)", "Código do Modelo (55=NF-e, 65=NFC-e, 28=CTe, 29=CTe complementar)"),
            ("ser", "VARCHAR(3)", "Série do documento"),
            ("num_doc", "VARCHAR(9)", "Número do documento fiscal"),
            ("chave_acesso", "VARCHAR(44)", "Chave de Acesso de 44 dígitos (NFe/CTe)"),
            ("cod_sit", "VARCHAR(2)", "Código de Situação (00=Regular, 01=Cancelado, 02=Denegado)"),
            ("status", "INTEGER", "Status de processamento no sistema"),
        ]
    },
    {
        "categoria": "4. Participante/Interveniente",
        "campos": [
            ("cod_part", "VARCHAR(60)", "Código do Participante (referência ao C150 - outros participantes)"),
        ]
    },
    {
        "categoria": "5. Identificação do Produto/Serviço",
        "campos": [
            ("codigo_produto", "VARCHAR(60)", "Código do Produto conforme tabela interna da empresa"),
            ("produto_id_anterior", "INTEGER", "ID anterior para histórico de conversão de produtos"),
            ("descricao_produto", "VARCHAR(200)", "Descrição do produto armazenada"),
            ("chave_produto", "VARCHAR(250)", "Chave única para identificação do produto (hash)"),
            ("num_item", "INTEGER", "Número sequencial do item no documento (1, 2, 3, ...)"),
        ]
    },
    {
        "categoria": "6. Quantidades e Unidades",
        "campos": [
            ("qtd", "DOUBLE PRECISION", "Quantidade do item (em unidade comercializada)"),
            ("unid", "VARCHAR(20)", "Unidade Comercializada (UN, KG, L, M, etc.)"),
            ("unid_ant", "VARCHAR(6)", "Unidade anterior (para compatibilidade)"),
            ("unid_convertida", "VARCHAR(10)", "Unidade convertida (para análise de conformidade)"),
            ("operador_conversao", "INTEGER", "Operador matemático para conversão (1=multiplicação, 2=divisão)"),
            ("fator_conversao", "DOUBLE PRECISION", "Fator de Conversão da unidade inventariada para comercializada"),
        ]
    },
    {
        "categoria": "7. Valores do Item",
        "campos": [
            ("vl_item", "DOUBLE PRECISION", "Valor Total do Item (sem desconto) = QTD × VUNIT"),
            ("vl_desc", "DOUBLE PRECISION", "Valor de Desconto do item (se houver)"),
            ("vl_ipi", "DOUBLE PRECISION", "Valor do IPI do item (quando aplicável)"),
            ("vl_abat_int", "DOUBLE PRECISION", "Valor de Abatimento não tributado"),
        ]
    },
    {
        "categoria": "8. ICMS - Tributação",
        "campos": [
            ("cst_icms", "VARCHAR(3)", "Código de Situação Tributária (00, 10, 20, 30, 40, 41, 50, 60, 70, 80, 90)"),
            ("cfop", "VARCHAR(4)", "Código Fiscal de Operação e Prestação (5100, 6100, 5102, etc.)"),
            ("vl_bc_icms", "DOUBLE PRECISION", "Valor da Base de Cálculo do ICMS"),
            ("aliq_icms", "DOUBLE PRECISION", "Alíquota do ICMS (em %) - Alíquota interna/interestadual"),
            ("aliq_nf", "DOUBLE PRECISION", "Alíquota da NF (conforme informado na nota)"),
            ("vl_icms", "DOUBLE PRECISION", "Valor do ICMS calculado"),
        ]
    },
    {
        "categoria": "9. ICMS ST (Substituição Tributária)",
        "campos": [
            ("vl_bc_icms_st", "DOUBLE PRECISION", "Valor da Base de Cálculo da ST"),
            ("vl_icms_st", "DOUBLE PRECISION", "Valor do ICMS ST retido/cobrado"),
        ]
    },
    {
        "categoria": "10. DIFAL (Diferencial de Alíquota)",
        "campos": [
            ("valor_difal", "DOUBLE PRECISION", "Valor do DIFAL (ICMS Diferencial entre UF origem/destino)"),
            ("aliq_difal_inter", "DOUBLE PRECISION", "Alíquota do DIFAL Interestadual (%)"),
        ]
    },
    {
        "categoria": "11. Finalidade e Classificação",
        "campos": [
            ("finalidade_mercadoria", "INTEGER", "Finalidade da Mercadoria (0=Revenda, 1=Consumo, 2=Ativo, 3=Uso/Consumo)"),
        ]
    },
    {
        "categoria": "12. Flags de Análise e Conformidade",
        "campos": [
            ("excluir_erro_aliquota", "BOOLEAN", "Flag para excluir item com erro de alíquota"),
            ("intimar_erro_aliquota", "BOOLEAN", "Flag para intimar erro de alíquota"),
            ("excluir_consumo", "BOOLEAN", "Excluir análise de consumo próprio"),
            ("intimar_consumo", "BOOLEAN", "Intimar para consumo próprio"),
            ("excluir_ativo", "BOOLEAN", "Excluir análise de ativo imobilizado"),
            ("intimar_ativo", "BOOLEAN", "Intimar para ativo imobilizado"),
            ("excluir_credito_st", "BOOLEAN", "Excluir análise de crédito ST"),
            ("intimar_credito_st", "BOOLEAN", "Intimar para crédito ST"),
            ("excluir_tranferencia_credito", "BOOLEAN", "Excluir análise de transferência de crédito"),
            ("intimar_tranferencia_credito", "BOOLEAN", "Intimar para transferência de crédito"),
            ("excluir_devolucao", "BOOLEAN", "Excluir análise de devolução"),
            ("intimar_devolucao", "BOOLEAN", "Intimar para devolução"),
            ("excluir_difal", "BOOLEAN", "Excluir análise de DIFAL"),
            ("intimar_difal", "BOOLEAN", "Intimar para DIFAL"),
            ("excluir_isento", "BOOLEAN", "Excluir operações isentas"),
            ("intimar_isento", "BOOLEAN", "Intimar operações isentas"),
            ("excluir_credito_sn", "BOOLEAN", "Excluir análise de crédito para Simples Nacional"),
            ("intimar_credito_sn", "BOOLEAN", "Intimar crédito para Simples Nacional"),
            ("excluir_cte_st", "BOOLEAN", "Excluir análise de CTe com ST"),
            ("intimar_cte_st", "BOOLEAN", "Intimar CTe com ST"),
            ("verificado_credito", "BOOLEAN", "Flag de verificação de crédito realizada"),
            ("excluir_calculo_estorno", "BOOLEAN", "Excluir do cálculo de estorno"),
            ("excluir_levantamento_estoque", "BOOLEAN", "Excluir do levantamento de estoque"),
        ]
    },
    {
        "categoria": "13. Informações de Devolução",
        "campos": [
            ("item_devolucao", "INTEGER", "Item do documento original em caso de devolução"),
            ("qtde_devolucao", "DOUBLE PRECISION", "Quantidade devolvida"),
            ("icms_devolucao", "DOUBLE PRECISION", "Valor de ICMS devolvido"),
        ]
    },
]

# Aba 1: Mapeamento Completo
ws_mapeamento = wb.create_sheet("Mapeamento Completo")
ws_mapeamento.column_dimensions['A'].width = 30
ws_mapeamento.column_dimensions['B'].width = 20
ws_mapeamento.column_dimensions['C'].width = 70

row = 1

# Cabeçalho principal
ws_mapeamento.merge_cells(f'A{row}:C{row}')
cell = ws_mapeamento[f'A{row}']
cell.value = "MAPEAMENTO DA TABELA EFD_REGC170 - ITENS DE SAÍDA EFD"
cell.font = Font(bold=True, size=14, color="FFFFFF")
cell.fill = header_fill
cell.alignment = center_align
row += 2

# Cabeçalhos de coluna
headers = ["Campo", "Tipo de Dado", "Descrição EFD"]
for col_num, header in enumerate(headers, 1):
    cell = ws_mapeamento.cell(row=row, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

row += 1

# Preencher dados por categoria
for categoria_data in campos:
    # Categoria
    ws_mapeamento.merge_cells(f'A{row}:C{row}')
    cell = ws_mapeamento[f'A{row}']
    cell.value = categoria_data['categoria']
    cell.font = category_font
    cell.fill = category_fill
    cell.alignment = left_align
    cell.border = thin_border
    row += 1
    
    # Campos da categoria
    for campo, tipo, descricao in categoria_data['campos']:
        cells = [
            ws_mapeamento.cell(row=row, column=1, value=campo),
            ws_mapeamento.cell(row=row, column=2, value=tipo),
            ws_mapeamento.cell(row=row, column=3, value=descricao),
        ]
        for cell in cells:
            cell.border = thin_border
            cell.alignment = left_align
        ws_mapeamento[f'B{row}'].alignment = center_align
        row += 1

# Aba 2: CST ICMS Referência
ws_cst = wb.create_sheet("CST ICMS")
ws_cst.column_dimensions['A'].width = 15
ws_cst.column_dimensions['B'].width = 60

cst_data = [
    ("CST", "Descrição"),
    ("00", "Operação Tributada"),
    ("10", "Operação Tributada com Substituição Tributária"),
    ("20", "Operação Isenta"),
    ("30", "Operação Isenta da Substituição Tributária"),
    ("40", "Operação com Suspensão do ICMS"),
    ("41", "Operação com Suspensão do ICMS para Consumidor Intermediário"),
    ("50", "Operação com Suspensão do ICMS para Consumidor Final"),
    ("60", "ICMS CT (CT-e)"),
    ("70", "Operação com Crédito de ICMS"),
    ("80", "Operação de Armazém"),
    ("90", "Outras"),
]

row = 1
ws_cst.merge_cells(f'A{row}:B{row}')
cell = ws_cst[f'A{row}']
cell.value = "CÓDIGOS DE SITUAÇÃO TRIBUTÁRIA (CST ICMS)"
cell.font = Font(bold=True, size=12, color="FFFFFF")
cell.fill = header_fill
cell.alignment = center_align
row += 2

for cst, desc in cst_data:
    cells = [
        ws_cst.cell(row=row, column=1, value=cst),
        ws_cst.cell(row=row, column=2, value=desc),
    ]
    if cst == "CST":
        for cell in cells:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
    else:
        ws_cst[f'A{row}'].alignment = center_align
        ws_cst[f'B{row}'].alignment = left_align
    
    for cell in cells:
        cell.border = thin_border
    row += 1

# Aba 3: CFOP Referência
ws_cfop = wb.create_sheet("CFOP")
ws_cfop.column_dimensions['A'].width = 15
ws_cfop.column_dimensions['B'].width = 70

cfop_data = [
    ("CFOP", "Descrição"),
    ("5100", "Venda de mercadoria adquirida ou recebida de terceiros"),
    ("5101", "Venda de mercadoria adquirida ou recebida de terceiros (CFOP específico)"),
    ("5102", "Venda de mercadoria adquirida ou recebida de terceiros (CFOP específico)"),
    ("5103", "Venda de mercadoria adquirida ou recebida de terceiros (CFOP específico)"),
    ("5104", "Venda de mercadoria adquirida ou recebida de terceiros (CFOP específico)"),
    ("5105", "Venda de mercadoria adquirida ou recebida de terceiros (CFOP específico)"),
    ("5109", "Outras saídas de mercadoria adquirida ou recebida de terceiros"),
    ("5201", "Devolução de compras"),
    ("5202", "Devolução de compras (especificado)"),
    ("5209", "Outras devoluções de compras"),
    ("5301", "Transferência de produção"),
    ("5302", "Transferência para industrialização"),
    ("5303", "Devolução de transferência"),
    ("5309", "Outras transferências"),
    ("6100", "Venda de mercadoria adquirida ou recebida de terceiros (ICMS não tributado)"),
    ("6101", "Venda para Consumidor Final"),
    ("6102", "Venda para Armazém"),
    ("6109", "Outras saídas"),
    ("6201", "Devolução (ICMS não tributado)"),
    ("6209", "Outras devoluções"),
    ("6301", "Transferência (ICMS não tributado)"),
    ("6309", "Outras transferências"),
]

row = 1
ws_cfop.merge_cells(f'A{row}:B{row}')
cell = ws_cfop[f'A{row}']
cell.value = "CÓDIGOS FISCAIS DE OPERAÇÃO E PRESTAÇÃO (CFOP) - SAÍDAS"
cell.font = Font(bold=True, size=12, color="FFFFFF")
cell.fill = header_fill
cell.alignment = center_align
row += 2

for cfop, desc in cfop_data:
    cells = [
        ws_cfop.cell(row=row, column=1, value=cfop),
        ws_cfop.cell(row=row, column=2, value=desc),
    ]
    if cfop == "CFOP":
        for cell in cells:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
    else:
        ws_cfop[f'A{row}'].alignment = center_align
        ws_cfop[f'B{row}'].alignment = left_align
    
    for cell in cells:
        cell.border = thin_border
    row += 1

# Aba 4: Informações Importantes
ws_info = wb.create_sheet("Informações")
ws_info.column_dimensions['A'].width = 80

row = 1
info_sections = [
    ("REGISTRO C170 NA EFD", [
        "• Obrigatoriedade: Um registro C170 para cada item de saída registrado no C100",
        "• Localização: Segue imediatamente após o registro C100 no arquivo EFD",
        "• Quantidade: Pode haver múltiplos registros C170 por C100 (um para cada item)",
        "• Relacionamento: Cada C170 deve estar vinculado a um C100 válido",
    ]),
    ("ANÁLISES TÍPICAS COM C170", [
        "• Auditoria de Alíquotas: Verificar se alíquota informada (aliq_nf) vs alíquota tabela (aliq_icms)",
        "• Análise de Crédito: Validar direito de crédito conforme CST e finalidade",
        "• Conferência de Estoque: Comparar quantidades com movimentação física",
        "• DIFAL: Validar operações interestaduais para ICMS Diferencial",
        "• Devolução: Rastrear itens devolvidos e estornos de crédito",
        "• Substituto Tributário: Validar retenção de ST em operações interestaduais",
    ]),
    ("FÓRMULAS IMPORTANTES", [
        "• vl_item = qtd × vl_unitário",
        "• vl_bc_icms = vl_item - vl_desc (base de cálculo)",
        "• vl_icms = vl_bc_icms × (aliq_icms / 100)",
        "• vl_bc_icms_st = vl_item (ou diferença se houver MVA)",
        "• vl_icms_st = vl_bc_icms_st × (aliq_st / 100)",
    ]),
]

for section_title, items in info_sections:
    ws_info.merge_cells(f'A{row}:A{row}')
    cell = ws_info[f'A{row}']
    cell.value = section_title
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = left_align
    row += 1
    
    for item in items:
        ws_info.merge_cells(f'A{row}:A{row}')
        cell = ws_info[f'A{row}']
        cell.value = item
        cell.alignment = left_align
        cell.border = thin_border
        ws_info.row_dimensions[row].height = 30
        row += 1
    
    row += 1

# Salvar arquivo
output_file = r"c:\Users\03002693901\OneDrive - SECRETARIA DE ESTADO DE FINANCAS\GEFIS_ENIO\consultas_oracle_sql_developer\consultas_sql\MAPEAMENTO_EFD_REGC170.xlsx"
wb.save(output_file)
print(f"✅ Arquivo Excel criado com sucesso: {output_file}")
print(f"\n📊 Abas criadas:")
print(f"   1. Mapeamento Completo - Todos os 47 campos com descrições")
print(f"   2. CST ICMS - Referência de códigos de situação tributária")
print(f"   3. CFOP - Referência de códigos fiscais de operação")
print(f"   4. Informações - Dicas e fórmulas importantes")
